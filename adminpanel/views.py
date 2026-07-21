from decimal import Decimal
from django.shortcuts import render, redirect
from django.db import models
from django.db.models import Sum, Q
from django.utils import timezone
from django.core.paginator import Paginator
from django.contrib import messages
from accounts.models import User
from receptionist.models import (
    Patient, OPDVisit, HospitalSettings, IPDAdmission,
    WardMaster, RoomMaster, BedMaster, IPDChargeMaster, IPDBill
)
from lab.models import (
    LaboratoryBill, LaboratoryBillItem, LaboratoryReport,
    LaboratoryReportResult, LaboratoryCase, LabTest, LabTestParameter
)


def format_indian_currency(amount):
    """
    Formats a numeric amount into Indian currency format.
    Example: 1245680 -> "₹ 12,45,680"
             46210   -> "₹ 46,210"
             0       -> "₹ 0"
    """
    try:
        amount_int = int(round(float(amount)))
    except (ValueError, TypeError):
        return "₹ 0"

    s = str(amount_int)
    if len(s) <= 3:
        formatted = s
    else:
        last3 = s[-3:]
        rest = s[:-3]
        groups = []
        while len(rest) > 2:
            groups.append(rest[-2:])
            rest = rest[:-2]
        if rest:
            groups.append(rest)
        groups.reverse()
        formatted = ",".join(groups) + "," + last3
    return f"₹ {formatted}"


def dashboard(request):
    today = timezone.localdate()

    # 1. Total Patients
    total_patients = Patient.objects.filter(is_active=True).count()

    # 2. Today's OPD Visits
    todays_opd = OPDVisit.objects.filter(
        visit_date=today
    ).exclude(
        status=OPDVisit.StatusChoices.CANCELLED
    ).count()

    # 3. Today's IPD Admitted
    todays_ipd_admitted = IPDAdmission.objects.filter(
        admission_date=today
    ).count()

    # 4. Today's Lab Requested
    todays_lab_requested = LaboratoryCase.objects.filter(
        created_at__date=today
    ).count()

    # 5. Occupied Beds & Available Wards
    occupied_beds_count = IPDAdmission.objects.filter(status='Admitted').count()
    if occupied_beds_count == 0 and BedMaster.objects.filter(is_occupied=True).exists():
        occupied_beds_count = BedMaster.objects.filter(is_occupied=True).count()

    available_wards_count = WardMaster.objects.count()

    # 6. Discharged Today
    discharged_today = IPDAdmission.objects.filter(
        discharge_date=today,
        status__in=['Discharged', 'Closed']
    ).count()

    # 7. Pending Laboratory Reports
    completed_visit_ids = LaboratoryReport.objects.filter(
        status__in=['COMPLETED', 'SENT']
    ).exclude(
        visit__laboratory_reports__status__in=['PENDING', 'IN_PROGRESS']
    ).values_list('visit_id', flat=True)

    pending_lab_reports_count = OPDVisit.objects.filter(
        Q(status=OPDVisit.StatusChoices.PENDING_LAB) |
        Q(laboratory_reports__status__in=['PENDING', 'IN_PROGRESS'])
    ).exclude(
        id__in=completed_visit_ids
    ).distinct().count()

    # --- 6 NEW REVENUE CARDS CALCULATIONS ---
    hospital = HospitalSettings.objects.first()
    consultation_fee = hospital.consultation_fee if (hospital and hospital.consultation_fee) else Decimal('200.00')

    # Card 1: Total OPD Revenue
    opd_visits_count = OPDVisit.objects.filter(
        visit_type=OPDVisit.VisitTypeChoices.NEW_VISIT
    ).exclude(
        status=OPDVisit.StatusChoices.CANCELLED
    ).count()
    total_opd_revenue_val = Decimal(opd_visits_count) * Decimal(str(consultation_fee))
    total_opd_revenue_formatted = format_indian_currency(total_opd_revenue_val)

    # Card 2: Today's OPD Revenue
    todays_opd_new_count = OPDVisit.objects.filter(
        visit_date=today,
        visit_type=OPDVisit.VisitTypeChoices.NEW_VISIT
    ).exclude(
        status=OPDVisit.StatusChoices.CANCELLED
    ).count()
    todays_opd_revenue_val = Decimal(todays_opd_new_count) * Decimal(str(consultation_fee))
    todays_opd_revenue_formatted = format_indian_currency(todays_opd_revenue_val)

    # Card 3: Total Laboratory Revenue
    total_lab_revenue_val = LaboratoryBill.objects.aggregate(
        total=Sum('grand_total')
    )['total'] or Decimal('0.00')
    total_lab_revenue_formatted = format_indian_currency(total_lab_revenue_val)

    # Card 4: Today's Laboratory Revenue
    todays_lab_revenue_val = LaboratoryBill.objects.filter(
        Q(created_at__date=today) | Q(bill_date=today)
    ).aggregate(
        total=Sum('grand_total')
    )['total'] or Decimal('0.00')
    todays_lab_revenue_formatted = format_indian_currency(todays_lab_revenue_val)

    # Card 5: Total IPD Revenue
    total_ipd_revenue_val = IPDBill.objects.filter(
        status__in=[IPDBill.StatusChoices.FINAL, IPDBill.StatusChoices.PAID]
    ).aggregate(
        total=Sum('net_amount')
    )['total'] or Decimal('0.00')
    total_ipd_revenue_formatted = format_indian_currency(total_ipd_revenue_val)

    # Card 6: Today's IPD Revenue
    todays_ipd_revenue_val = IPDBill.objects.filter(
        Q(created_at__date=today) | Q(bill_date=today),
        status__in=[IPDBill.StatusChoices.FINAL, IPDBill.StatusChoices.PAID]
    ).aggregate(
        total=Sum('net_amount')
    )['total'] or Decimal('0.00')
    todays_ipd_revenue_formatted = format_indian_currency(todays_ipd_revenue_val)

    context = {
        'total_patients': f"{total_patients:,}",
        'todays_opd': f"{todays_opd:,}",
        'todays_ipd_admitted': f"{todays_ipd_admitted:,}",
        'todays_lab_requested': f"{todays_lab_requested:,}",
        'available_wards': f"{available_wards_count:,}",
        'occupied_beds': f"{occupied_beds_count:,}",
        'discharged_today': f"{discharged_today:,}",
        'pending_lab_reports': f"{pending_lab_reports_count:,}",
        
        # 6 New Revenue Cards
        'total_opd_revenue': total_opd_revenue_formatted,
        'todays_opd_revenue': todays_opd_revenue_formatted,
        'total_lab_revenue': total_lab_revenue_formatted,
        'todays_lab_revenue': todays_lab_revenue_formatted,
        'total_ipd_revenue': total_ipd_revenue_formatted,
        'todays_ipd_revenue': todays_ipd_revenue_formatted,
    }
    return render(request, "adminpanel/dashboard.html", context)


def users_admin(request):
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add":
            email = request.POST.get("email", "").strip()
            first_name = request.POST.get("first_name", "").strip()
            last_name = request.POST.get("last_name", "").strip()
            phone_number = request.POST.get("phone_number", "").strip()
            role = request.POST.get("role", "").strip()
            password = request.POST.get("password", "").strip()

            if email and password and role:
                if User.objects.filter(email=email).exists():
                    messages.error(request, f"User with email '{email}' already exists.")
                else:
                    base_username = email.split("@")[0]
                    username = base_username
                    counter = 1
                    while User.objects.filter(username=username).exists():
                        username = f"{base_username}{counter}"
                        counter += 1

                    is_staff_user = (role == User.Role.ADMIN)
                    user = User(
                        email=email,
                        username=username,
                        first_name=first_name,
                        last_name=last_name,
                        phone_number=phone_number,
                        role=role,
                        is_active=True,
                        is_staff=is_staff_user,
                    )
                    user.set_password(password)
                    user.save()
                    messages.success(request, f"User '{email}' created successfully.")
            else:
                messages.error(request, "Email, password, and role are required.")
            return redirect("adminpanel:users_admin")

        elif action == "edit":
            user_id = request.POST.get("user_id")
            try:
                user = User.objects.get(id=user_id)
                user.first_name = request.POST.get("first_name", user.first_name).strip()
                user.last_name = request.POST.get("last_name", user.last_name).strip()
                user.phone_number = request.POST.get("phone_number", user.phone_number).strip()
                new_role = request.POST.get("role", user.role).strip()
                if new_role in User.Role.values:
                    user.role = new_role

                new_email = request.POST.get("email", user.email).strip()
                if new_email and new_email != user.email:
                    if User.objects.filter(email=new_email).exclude(id=user.id).exists():
                        messages.error(request, f"Email '{new_email}' is already taken.")
                    else:
                        user.email = new_email

                new_password = request.POST.get("password", "").strip()
                if new_password:
                    user.set_password(new_password)

                user.is_staff = (user.role == User.Role.ADMIN)
                user.save()
                messages.success(request, f"User '{user.email}' updated successfully.")
            except User.DoesNotExist:
                messages.error(request, "User not found.")
            return redirect("adminpanel:users_admin")

        elif action == "delete":
            user_id = request.POST.get("user_id")
            try:
                user = User.objects.get(id=user_id)
                if user == request.user:
                    messages.error(request, "You cannot delete your own logged-in account.")
                else:
                    email = user.email
                    user.delete()
                    messages.success(request, f"User '{email}' deleted successfully.")
            except User.DoesNotExist:
                messages.error(request, "User not found.")
            return redirect("adminpanel:users_admin")

    # Metrics Summary Cards
    all_users = User.objects.all().order_by("-date_joined")
    total_users = all_users.count()
    total_doctors = all_users.filter(role=User.Role.DOCTOR).count()
    total_receptionists = all_users.filter(role=User.Role.RECEPTIONIST).count()
    total_lab_staff = all_users.filter(role=User.Role.LAB_ADMINISTRATOR).count()
    total_admins = all_users.filter(role=User.Role.ADMIN).count()

    # Search Filter
    search_query = request.GET.get("q", "").strip() or request.GET.get("search", "").strip()
    users_qs = all_users
    if search_query:
        users_qs = users_qs.filter(
            models.Q(first_name__icontains=search_query) |
            models.Q(last_name__icontains=search_query) |
            models.Q(email__icontains=search_query) |
            models.Q(phone_number__icontains=search_query) |
            models.Q(username__icontains=search_query)
        )

    # Role Filter
    selected_role = request.GET.get("role", "").strip()
    if selected_role in User.Role.values:
        users_qs = users_qs.filter(role=selected_role)

    # Pagination
    paginator = Paginator(users_qs, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "total_users": total_users,
        "total_doctors": total_doctors,
        "total_receptionists": total_receptionists,
        "total_lab_staff": total_lab_staff,
        "total_admins": total_admins,
        "users": page_obj,
        "page_obj": page_obj,
        "search_query": search_query,
        "selected_role": selected_role,
    }
    return render(request, "adminpanel/users_management.html", context)


import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse


def _get_filtered_patients_queryset(request):
    all_patients = Patient.objects.filter(is_active=True).order_by("-created_at")

    search_query = request.GET.get("q", "").strip() or request.GET.get("search", "").strip()
    patients_qs = all_patients
    if search_query:
        patients_qs = patients_qs.filter(
            Q(full_name__icontains=search_query) |
            Q(uhid__icontains=search_query) |
            Q(mobile_number__icontains=search_query)
        )

    selected_filter = request.GET.get("filter", "all").strip().lower()
    if selected_filter == "opd":
        patients_qs = patients_qs.filter(visits__isnull=False).distinct()
    elif selected_filter == "ipd":
        patients_qs = patients_qs.filter(ipd_admissions__isnull=False).distinct()
    elif selected_filter == "discharged":
        patients_qs = patients_qs.filter(
            Q(ipd_admissions__status__in=["Discharged", "Closed"]) |
            Q(ipd_admissions__discharge_date__isnull=False)
        ).distinct()

    return patients_qs.prefetch_related('visits', 'ipd_admissions'), search_query, selected_filter, all_patients


def _build_patient_export_row(p):
    latest_visit = p.visits.all().first()
    latest_ipd = p.ipd_admissions.all().first()

    visit_type = latest_visit.visit_type if latest_visit else "New Visit"

    if latest_ipd:
        patient_type = "IPD"
        is_discharged = (latest_ipd.status in ["Discharged", "Closed"]) or (latest_ipd.discharge_date is not None)
        status_text = "Discharged" if is_discharged else "Admitted"
    elif latest_visit:
        patient_type = "OPD"
        status_text = "Active"
    else:
        patient_type = "OPD"
        status_text = "Active"

    return {
        'uhid': p.uhid or "",
        'full_name': p.full_name or "",
        'age': str(p.age) if p.age is not None else "-",
        'gender': p.gender or "-",
        'mobile_number': p.mobile_number or "-",
        'visit_type': visit_type,
        'patient_type': patient_type,
        'created_at': p.created_at.strftime("%d %b %Y") if p.created_at else "-",
        'status': status_text,
    }


def patients_all(request):
    patients_qs, search_query, selected_filter, all_patients = _get_filtered_patients_queryset(request)

    # Dynamic metrics summary
    total_patients_count = all_patients.count()
    new_patients_count = Patient.objects.filter(
        is_active=True,
        visits__visit_type=OPDVisit.VisitTypeChoices.NEW_VISIT
    ).distinct().count()
    opd_patients_count = Patient.objects.filter(
        is_active=True,
        visits__isnull=False
    ).distinct().count()
    ipd_patients_count = Patient.objects.filter(
        is_active=True,
        ipd_admissions__isnull=False
    ).distinct().count()

    # Pagination
    paginator = Paginator(patients_qs, 15)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Build row context objects
    patients_data = []
    for p in page_obj:
        latest_visit = p.visits.all().first()
        latest_ipd = p.ipd_admissions.all().first()

        visit_type = latest_visit.visit_type if latest_visit else "New Visit"

        if latest_ipd:
            patient_type = "IPD"
            is_discharged = (latest_ipd.status in ["Discharged", "Closed"]) or (latest_ipd.discharge_date is not None)
            if is_discharged:
                status_text = "Discharged"
                status_badge_class = "badge bg-secondary-subtle text-secondary border border-secondary-subtle fw-medium px-2 py-1"
                status_display = "⚪ Discharged"
            else:
                status_text = "Admitted"
                status_badge_class = "badge bg-primary-subtle text-primary border border-primary-subtle fw-medium px-2 py-1"
                status_display = "🔵 Admitted"
        elif latest_visit:
            patient_type = "OPD"
            status_text = "Active"
            status_badge_class = "badge bg-success-subtle text-success border border-success-subtle fw-medium px-2 py-1"
            status_display = "🟢 Active"
        else:
            patient_type = "OPD"
            status_text = "Active"
            status_badge_class = "badge bg-success-subtle text-success border border-success-subtle fw-medium px-2 py-1"
            status_display = "🟢 Active"

        patients_data.append({
            'patient': p,
            'uhid': p.uhid,
            'full_name': p.full_name,
            'age': p.age,
            'gender': p.gender,
            'mobile_number': p.mobile_number,
            'visit_type': visit_type,
            'patient_type': patient_type,
            'created_at': p.created_at,
            'status_text': status_text,
            'status_badge_class': status_badge_class,
            'status_display': status_display,
        })

    context = {
        'total_patients': f"{total_patients_count:,}",
        'new_patients': f"{new_patients_count:,}",
        'opd_patients': f"{opd_patients_count:,}",
        'ipd_patients': f"{ipd_patients_count:,}",
        'patients_data': patients_data,
        'page_obj': page_obj,
        'search_query': search_query,
        'selected_filter': selected_filter,
    }
    return render(request, "adminpanel/all_patients.html", context)


def export_patients_csv(request):
    patients_qs, _, selected_filter, _ = _get_filtered_patients_queryset(request)

    filename_suffix = f"_{selected_filter}" if selected_filter and selected_filter != "all" else ""
    filename = f"all_patients_report{filename_suffix}.csv"

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        "UHID / Patient ID",
        "Patient Name",
        "Age",
        "Gender",
        "Mobile Number",
        "Visit Type",
        "Patient Type",
        "Registration Date",
        "Status"
    ])

    for p in patients_qs:
        row = _build_patient_export_row(p)
        writer.writerow([
            row['uhid'],
            row['full_name'],
            row['age'],
            row['gender'],
            row['mobile_number'],
            row['visit_type'],
            row['patient_type'],
            row['created_at'],
            row['status']
        ])

    return response


def export_patients_excel(request):
    patients_qs, _, selected_filter, _ = _get_filtered_patients_queryset(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients Report"

    headers = [
        "UHID / Patient ID",
        "Patient Name",
        "Age",
        "Gender",
        "Mobile Number",
        "Visit Type",
        "Patient Type",
        "Registration Date",
        "Status"
    ]

    ws.append(headers)

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for p in patients_qs:
        row = _build_patient_export_row(p)
        row_data = [
            row['uhid'],
            row['full_name'],
            row['age'],
            row['gender'],
            row['mobile_number'],
            row['visit_type'],
            row['patient_type'],
            row['created_at'],
            row['status']
        ]
        ws.append(row_data)

        row_idx = ws.max_row
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            if col_num in [1, 3, 4, 5, 6, 7, 8, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    filename_suffix = f"_{selected_filter}" if selected_filter and selected_filter != "all" else ""
    filename = f"all_patients_report{filename_suffix}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def ipd_master(request):
    from django.db.models import Count, ProtectedError
    from django.urls import reverse

    if request.method == "POST":
        action = request.POST.get("action", "").strip()

        if action == "add_ward":
            ward_name = request.POST.get("name", "").strip()
            description = request.POST.get("description", "").strip()

            if not ward_name:
                messages.error(request, "Ward name is required.")
            elif WardMaster.objects.filter(name__iexact=ward_name).exists():
                messages.error(request, f"Ward with name '{ward_name}' already exists.")
            else:
                WardMaster.objects.create(name=ward_name, description=description)
                messages.success(request, f"Ward '{ward_name}' created successfully.")
            return redirect(f"{reverse('adminpanel:ipd_master')}?tab=wards")

        elif action == "add_room":
            ward_id = request.POST.get("ward_id")
            room_number = request.POST.get("room_number", "").strip()
            room_type = request.POST.get("room_type", "Room").strip()
            capacity = request.POST.get("capacity", 1)

            try:
                capacity = int(capacity)
            except (ValueError, TypeError):
                capacity = 1

            if not ward_id or not room_number:
                messages.error(request, "Ward and Room Number are required.")
            else:
                try:
                    ward = WardMaster.objects.get(id=ward_id)
                    if RoomMaster.objects.filter(ward=ward, room_number__iexact=room_number).exists():
                        messages.error(request, f"Room '{room_number}' already exists in Ward '{ward.name}'.")
                    else:
                        RoomMaster.objects.create(
                            ward=ward,
                            room_number=room_number,
                            room_type=room_type,
                            capacity=capacity
                        )
                        messages.success(request, f"Room '{room_number}' added to Ward '{ward.name}'.")
                except WardMaster.DoesNotExist:
                    messages.error(request, "Selected Ward does not exist.")
            return redirect(f"{reverse('adminpanel:ipd_master')}?tab=rooms")

        elif action == "add_bed":
            room_id = request.POST.get("room_id")
            bed_number = request.POST.get("bed_number", "").strip()

            if not room_id or not bed_number:
                messages.error(request, "Room and Bed Number are required.")
            else:
                try:
                    room = RoomMaster.objects.select_related('ward').get(id=room_id)
                    if BedMaster.objects.filter(room=room, bed_number__iexact=bed_number).exists():
                        messages.error(request, f"Bed '{bed_number}' already exists in {room}.")
                    else:
                        BedMaster.objects.create(
                            room=room,
                            bed_number=bed_number,
                            is_occupied=False
                        )
                        messages.success(request, f"Bed '{bed_number}' added to {room}.")
                except RoomMaster.DoesNotExist:
                    messages.error(request, "Selected Room does not exist.")
            return redirect(f"{reverse('adminpanel:ipd_master')}?tab=beds")

        elif action == "update_charge":
            charge_id = request.POST.get("charge_id")
            amount = request.POST.get("amount", "").strip()
            is_active_val = request.POST.get("is_active")

            if not charge_id or not amount:
                messages.error(request, "Charge selection and rate/amount are required.")
            else:
                try:
                    amount_dec = Decimal(amount)
                    charge = IPDChargeMaster.objects.get(id=charge_id)
                    charge.amount = amount_dec
                    if is_active_val is not None:
                        charge.is_active = (str(is_active_val).lower() in ['true', '1', 'on', 'yes'])
                    charge.save()
                    messages.success(request, f"IPD Charge '{charge.name}' updated successfully to ₹{charge.amount}.")
                except (ValueError, TypeError, Decimal.InvalidOperation):
                    messages.error(request, "Invalid amount entered.")
                except IPDChargeMaster.DoesNotExist:
                    messages.error(request, "Selected IPD Charge record does not exist.")
            return redirect(f"{reverse('adminpanel:ipd_master')}?tab=charges")

        elif action == "delete_ward":
            ward_id = request.POST.get("ward_id")
            try:
                ward = WardMaster.objects.get(id=ward_id)
                ward_name = ward.name
                ward.delete()
                messages.success(request, f"Ward '{ward_name}' deleted successfully.")
            except ProtectedError:
                messages.error(request, "Cannot delete ward because active rooms or admissions are linked to it.")
            except Exception as e:
                messages.error(request, f"Error deleting ward: {str(e)}")
            return redirect(f"{reverse('adminpanel:ipd_master')}?tab=wards")

        elif action == "delete_room":
            room_id = request.POST.get("room_id")
            try:
                room = RoomMaster.objects.get(id=room_id)
                room_info = str(room)
                room.delete()
                messages.success(request, f"Room '{room_info}' deleted successfully.")
            except ProtectedError:
                messages.error(request, "Cannot delete room because active beds or admissions are linked to it.")
            except Exception as e:
                messages.error(request, f"Error deleting room: {str(e)}")
            return redirect(f"{reverse('adminpanel:ipd_master')}?tab=rooms")

        elif action == "delete_bed":
            bed_id = request.POST.get("bed_id")
            try:
                bed = BedMaster.objects.get(id=bed_id)
                if bed.is_occupied:
                    messages.error(request, f"Cannot delete Bed '{bed.bed_number}' as it is currently occupied.")
                else:
                    bed_info = str(bed)
                    bed.delete()
                    messages.success(request, f"Bed '{bed_info}' deleted successfully.")
            except ProtectedError:
                messages.error(request, "Cannot delete bed because active admissions are linked to it.")
            except Exception as e:
                messages.error(request, f"Error deleting bed: {str(e)}")
            return redirect(f"{reverse('adminpanel:ipd_master')}?tab=beds")

    # Summary Cards
    total_wards_count = WardMaster.objects.count()
    total_rooms_count = RoomMaster.objects.count()
    total_beds_count = BedMaster.objects.count()

    # Search & Tab Parameters
    search_query = request.GET.get("q", "").strip() or request.GET.get("search", "").strip()
    active_tab = request.GET.get("tab", "wards").strip().lower()
    if active_tab not in ["wards", "rooms", "beds", "charges"]:
        active_tab = "wards"

    # Wards Queryset
    wards_qs = WardMaster.objects.annotate(total_rooms=Count('rooms')).order_by('name')
    if search_query and active_tab == "wards":
        wards_qs = wards_qs.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    wards_paginator = Paginator(wards_qs, 10)
    wards_page = wards_paginator.get_page(request.GET.get("page") if active_tab == "wards" else 1)

    # Rooms Queryset
    rooms_qs = RoomMaster.objects.select_related('ward').annotate(total_beds=Count('beds')).order_by('ward__name', 'room_number')
    if search_query and active_tab == "rooms":
        rooms_qs = rooms_qs.filter(
            Q(room_number__icontains=search_query) |
            Q(ward__name__icontains=search_query) |
            Q(room_type__icontains=search_query)
        )
    rooms_paginator = Paginator(rooms_qs, 10)
    rooms_page = rooms_paginator.get_page(request.GET.get("page") if active_tab == "rooms" else 1)

    # Beds Queryset
    beds_qs = BedMaster.objects.select_related('room__ward').order_by('room__ward__name', 'room__room_number', 'bed_number')
    if search_query and active_tab == "beds":
        beds_qs = beds_qs.filter(
            Q(bed_number__icontains=search_query) |
            Q(room__room_number__icontains=search_query) |
            Q(room__ward__name__icontains=search_query)
        )
    beds_paginator = Paginator(beds_qs, 10)
    beds_page = beds_paginator.get_page(request.GET.get("page") if active_tab == "beds" else 1)

    # Charges Queryset
    charges_qs = IPDChargeMaster.objects.all().order_by('name')
    if search_query and active_tab == "charges":
        charges_qs = charges_qs.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(ward__icontains=search_query) |
            Q(charge_type__icontains=search_query)
        )
    charges_paginator = Paginator(charges_qs, 10)
    charges_page = charges_paginator.get_page(request.GET.get("page") if active_tab == "charges" else 1)

    # Master lists for Modals dropdowns
    all_wards_list = WardMaster.objects.all().order_by('name')
    all_rooms_list = RoomMaster.objects.select_related('ward').all().order_by('ward__name', 'room_number')
    all_charges_list = IPDChargeMaster.objects.all().order_by('name')

    context = {
        'total_wards': f"{total_wards_count:,}",
        'total_rooms': f"{total_rooms_count:,}",
        'total_beds': f"{total_beds_count:,}",
        'active_tab': active_tab,
        'search_query': search_query,
        'wards_page': wards_page,
        'rooms_page': rooms_page,
        'beds_page': beds_page,
        'charges_page': charges_page,
        'all_wards_list': all_wards_list,
        'all_rooms_list': all_rooms_list,
        'all_charges_list': all_charges_list,
    }
    return render(request, "adminpanel/ipd_master.html", context)


def lab_master(request):
    from django.db.models import Count, ProtectedError
    from django.urls import reverse

    if request.method == "POST":
        action = request.POST.get("action", "").strip()

        if action == "add_test":
            name = request.POST.get("name", "").strip()
            price_str = request.POST.get("price", "").strip()
            is_active_val = request.POST.get("is_active", "true")

            if not name or not price_str:
                messages.error(request, "Test name and price are required.")
            elif LabTest.objects.filter(name__iexact=name).exists():
                messages.error(request, f"A Lab Test named '{name}' already exists.")
            else:
                try:
                    price = Decimal(price_str)
                    is_active = (str(is_active_val).lower() in ['true', '1', 'on', 'yes'])
                    test = LabTest.objects.create(name=name, price=price, is_active=is_active)
                    messages.success(request, f"Lab Test '{test.name}' added successfully.")
                except (ValueError, TypeError, Decimal.InvalidOperation):
                    messages.error(request, "Invalid price entered.")
                except Exception as e:
                    messages.error(request, f"Error adding Lab Test: {str(e)}")
            return redirect(f"{reverse('adminpanel:lab_master')}?tab=tests")

        elif action == "edit_test":
            test_id = request.POST.get("test_id")
            name = request.POST.get("name", "").strip()
            price_str = request.POST.get("price", "").strip()
            is_active_val = request.POST.get("is_active", "true")

            if not test_id or not name or not price_str:
                messages.error(request, "Test selection, name, and price are required.")
            elif LabTest.objects.filter(name__iexact=name).exclude(id=test_id).exists():
                messages.error(request, f"Another Lab Test named '{name}' already exists.")
            else:
                try:
                    price = Decimal(price_str)
                    test = LabTest.objects.get(id=test_id)
                    test.name = name
                    test.price = price
                    if is_active_val is not None:
                        test.is_active = (str(is_active_val).lower() in ['true', '1', 'on', 'yes'])
                    test.save()
                    messages.success(request, f"Lab Test '{test.name}' updated successfully.")
                except (ValueError, TypeError, Decimal.InvalidOperation):
                    messages.error(request, "Invalid price entered.")
                except LabTest.DoesNotExist:
                    messages.error(request, "Selected Lab Test does not exist.")
                except Exception as e:
                    messages.error(request, f"Error updating Lab Test: {str(e)}")
            return redirect(f"{reverse('adminpanel:lab_master')}?tab=tests")

        elif action == "delete_test":
            test_id = request.POST.get("test_id")
            try:
                test = LabTest.objects.get(id=test_id)
                test_name = test.name

                if LaboratoryBillItem.objects.filter(test=test).exists():
                    messages.error(request, f"Cannot delete Lab Test '{test_name}' as it is referenced in existing laboratory bills.")
                elif LaboratoryReport.objects.filter(lab_test=test).exists():
                    messages.error(request, f"Cannot delete Lab Test '{test_name}' as it is referenced in existing laboratory reports.")
                else:
                    test.delete()
                    messages.success(request, f"Lab Test '{test_name}' deleted successfully.")
            except ProtectedError:
                messages.error(request, "Cannot delete Lab Test as it is protected by related historical records.")
            except LabTest.DoesNotExist:
                messages.error(request, "Selected Lab Test does not exist.")
            except Exception as e:
                messages.error(request, f"Error deleting Lab Test: {str(e)}")
            return redirect(f"{reverse('adminpanel:lab_master')}?tab=tests")

        elif action == "add_parameter":
            lab_test_id = request.POST.get("lab_test_id")
            parameter_name = request.POST.get("parameter_name", "").strip()
            unit = request.POST.get("unit", "").strip()
            range_type = request.POST.get("range_type", "common")
            common_ref = request.POST.get("common_reference_range", "").strip()
            male_ref = request.POST.get("male_reference_range", "").strip()
            female_ref = request.POST.get("female_reference_range", "").strip()
            display_order_str = request.POST.get("display_order", "0").strip()
            is_active_val = request.POST.get("is_active", "true")

            if not lab_test_id or not parameter_name:
                messages.error(request, "Lab Test selection and Parameter Name are required.")
            elif LabTestParameter.objects.filter(lab_test_id=lab_test_id, parameter_name__iexact=parameter_name).exists():
                messages.error(request, f"Parameter '{parameter_name}' already exists for this Lab Test.")
            else:
                try:
                    lab_test = LabTest.objects.get(id=lab_test_id)
                    display_order = int(display_order_str) if display_order_str.isdigit() else 0
                    is_active = (str(is_active_val).lower() in ['true', '1', 'on', 'yes'])

                    if range_type == "gender" and male_ref and female_ref:
                        common_ref = ""
                    else:
                        male_ref = ""
                        female_ref = ""
                        if not common_ref:
                            common_ref = "Standard"

                    param = LabTestParameter(
                        lab_test=lab_test,
                        parameter_name=parameter_name,
                        unit=unit,
                        common_reference_range=common_ref,
                        male_reference_range=male_ref,
                        female_reference_range=female_ref,
                        display_order=display_order,
                        is_active=is_active
                    )
                    param.save()
                    messages.success(request, f"Parameter '{parameter_name}' added to '{lab_test.name}' successfully.")
                except LabTest.DoesNotExist:
                    messages.error(request, "Selected Lab Test does not exist.")
                except Exception as e:
                    messages.error(request, f"Error adding Test Parameter: {str(e)}")
            return redirect(f"{reverse('adminpanel:lab_master')}?tab=parameters")

        elif action == "edit_parameter":
            param_id = request.POST.get("param_id")
            lab_test_id = request.POST.get("lab_test_id")
            parameter_name = request.POST.get("parameter_name", "").strip()
            unit = request.POST.get("unit", "").strip()
            range_type = request.POST.get("range_type", "common")
            common_ref = request.POST.get("common_reference_range", "").strip()
            male_ref = request.POST.get("male_reference_range", "").strip()
            female_ref = request.POST.get("female_reference_range", "").strip()
            display_order_str = request.POST.get("display_order", "0").strip()
            is_active_val = request.POST.get("is_active", "true")

            if not param_id or not parameter_name:
                messages.error(request, "Parameter selection and Parameter Name are required.")
            else:
                try:
                    param = LabTestParameter.objects.get(id=param_id)
                    if lab_test_id:
                        lab_test = LabTest.objects.get(id=lab_test_id)
                        param.lab_test = lab_test

                    if LabTestParameter.objects.filter(lab_test=param.lab_test, parameter_name__iexact=parameter_name).exclude(id=param_id).exists():
                        messages.error(request, f"Another parameter named '{parameter_name}' already exists for test '{param.lab_test.name}'.")
                    else:
                        param.parameter_name = parameter_name
                        param.unit = unit
                        param.display_order = int(display_order_str) if display_order_str.isdigit() else 0
                        if is_active_val is not None:
                            param.is_active = (str(is_active_val).lower() in ['true', '1', 'on', 'yes'])

                        if range_type == "gender" and male_ref and female_ref:
                            param.male_reference_range = male_ref
                            param.female_reference_range = female_ref
                            param.common_reference_range = ""
                        else:
                            param.male_reference_range = ""
                            param.female_reference_range = ""
                            param.common_reference_range = common_ref if common_ref else "Standard"

                        param.save()
                        messages.success(request, f"Parameter '{param.parameter_name}' updated successfully.")
                except LabTestParameter.DoesNotExist:
                    messages.error(request, "Selected Test Parameter does not exist.")
                except LabTest.DoesNotExist:
                    messages.error(request, "Selected Lab Test does not exist.")
                except Exception as e:
                    messages.error(request, f"Error updating Test Parameter: {str(e)}")
            return redirect(f"{reverse('adminpanel:lab_master')}?tab=parameters")

        elif action == "delete_parameter":
            param_id = request.POST.get("param_id")
            try:
                param = LabTestParameter.objects.get(id=param_id)
                param_name = param.parameter_name

                if LaboratoryReportResult.objects.filter(parameter=param).exists():
                    messages.error(request, f"Cannot delete parameter '{param_name}' as it is referenced in existing laboratory report results.")
                else:
                    param.delete()
                    messages.success(request, f"Parameter '{param_name}' deleted successfully.")
            except ProtectedError:
                messages.error(request, "Cannot delete parameter as it is protected by related historical records.")
            except LabTestParameter.DoesNotExist:
                messages.error(request, "Selected Test Parameter does not exist.")
            except Exception as e:
                messages.error(request, f"Error deleting Test Parameter: {str(e)}")
            return redirect(f"{reverse('adminpanel:lab_master')}?tab=parameters")

    # Summary Cards
    total_lab_tests_count = LabTest.objects.count()
    total_test_parameters_count = LabTestParameter.objects.count()
    all_lab_reports_count = LaboratoryReport.objects.count()

    # Search & Tab Parameters
    search_query = request.GET.get("q", "").strip() or request.GET.get("search", "").strip()
    active_tab = request.GET.get("tab", "tests").strip().lower()
    if active_tab not in ["tests", "parameters", "reports"]:
        active_tab = "tests"

    # Laboratory Tests Queryset
    tests_qs = LabTest.objects.annotate(param_count=Count('parameters')).order_by('name')
    if search_query and active_tab == "tests":
        tests_qs = tests_qs.filter(
            Q(name__icontains=search_query)
        )
    tests_paginator = Paginator(tests_qs, 10)
    tests_page = tests_paginator.get_page(request.GET.get("page") if active_tab == "tests" else 1)

    # Test Parameters Queryset
    params_qs = LabTestParameter.objects.select_related('lab_test').order_by('lab_test__name', 'display_order', 'parameter_name')
    if search_query and active_tab == "parameters":
        params_qs = params_qs.filter(
            Q(parameter_name__icontains=search_query) |
            Q(lab_test__name__icontains=search_query) |
            Q(unit__icontains=search_query)
        )
    params_paginator = Paginator(params_qs, 10)
    params_page = params_paginator.get_page(request.GET.get("page") if active_tab == "parameters" else 1)

    # All Lab Reports Queryset
    reports_qs = LaboratoryReport.objects.select_related('patient', 'visit', 'visit__created_by', 'lab_test', 'generated_by').order_by('-generated_date')
    if search_query and active_tab == "reports":
        reports_qs = reports_qs.filter(
            Q(report_number__icontains=search_query) |
            Q(patient__full_name__icontains=search_query) |
            Q(patient__uhid__icontains=search_query) |
            Q(lab_test__name__icontains=search_query)
        )
    reports_paginator = Paginator(reports_qs, 10)
    reports_page = reports_paginator.get_page(request.GET.get("page") if active_tab == "reports" else 1)

    # Master lists for Modals dropdowns
    all_tests_list = LabTest.objects.all().order_by('name')

    context = {
        'total_lab_tests': f"{total_lab_tests_count:,}",
        'total_test_parameters': f"{total_test_parameters_count:,}",
        'all_lab_reports': f"{all_lab_reports_count:,}",
        'active_tab': active_tab,
        'search_query': search_query,
        'tests_page': tests_page,
        'params_page': params_page,
        'reports_page': reports_page,
        'all_tests_list': all_tests_list,
    }
    return render(request, "adminpanel/lab_master.html", context)


def reports_dashboard(request):
    from datetime import datetime
    import csv
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from decimal import Decimal

    # Summary Cards
    opd_reports_count = OPDVisit.objects.count()
    ipd_reports_count = IPDAdmission.objects.count()
    lab_reports_count = LaboratoryReport.objects.count()
    billing_reports_count = OPDVisit.objects.count() + IPDBill.objects.count() + LaboratoryBill.objects.count()

    # Search, Active Tab & Filters
    search_query = request.GET.get("q", "").strip() or request.GET.get("search", "").strip()
    active_tab = request.GET.get("tab", "opd").strip().lower()
    if active_tab not in ["opd", "ipd", "lab", "billing"]:
        active_tab = "opd"

    start_date_str = request.GET.get("start_date", "").strip()
    end_date_str = request.GET.get("end_date", "").strip()
    status_filter = request.GET.get("status", "").strip()
    patient_type_filter = request.GET.get("patient_type", "").strip()
    billing_type_filter = (
        request.GET.get("billing_type", "").strip()
        or request.GET.get("bill_type", "").strip()
        or request.GET.get("patient_type", "").strip()
        or "all"
    )
    export_format = request.GET.get("export", "").strip().lower()

    # Date Parsing
    start_date = None
    end_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    # 1. OPD Reports Queryset
    opd_qs = OPDVisit.objects.select_related('patient', 'created_by').order_by('-visit_date', '-visit_time')
    if search_query:
        opd_qs = opd_qs.filter(
            Q(opd_number__icontains=search_query) |
            Q(patient__full_name__icontains=search_query) |
            Q(patient__uhid__icontains=search_query) |
            Q(created_by__first_name__icontains=search_query) |
            Q(created_by__last_name__icontains=search_query) |
            Q(created_by__username__icontains=search_query)
        )
    if start_date:
        opd_qs = opd_qs.filter(visit_date__gte=start_date)
    if end_date:
        opd_qs = opd_qs.filter(visit_date__lte=end_date)
    if status_filter and status_filter.lower() != 'all':
        opd_qs = opd_qs.filter(status__iexact=status_filter)

    # 2. IPD Reports Queryset
    ipd_qs = IPDAdmission.objects.select_related('patient').order_by('-admission_date', '-admission_time')
    if search_query:
        ipd_qs = ipd_qs.filter(
            Q(receipt_number__icontains=search_query) |
            Q(patient__full_name__icontains=search_query) |
            Q(patient__uhid__icontains=search_query) |
            Q(ward_type__icontains=search_query)
        )
    if start_date:
        ipd_qs = ipd_qs.filter(admission_date__gte=start_date)
    if end_date:
        ipd_qs = ipd_qs.filter(admission_date__lte=end_date)
    if status_filter and status_filter.lower() != 'all':
        ipd_qs = ipd_qs.filter(status__iexact=status_filter)

    # 3. Laboratory Reports Queryset
    lab_qs = LaboratoryReport.objects.select_related('patient', 'visit', 'visit__created_by', 'lab_test').order_by('-generated_date')
    if search_query:
        lab_qs = lab_qs.filter(
            Q(report_number__icontains=search_query) |
            Q(patient__full_name__icontains=search_query) |
            Q(patient__uhid__icontains=search_query) |
            Q(lab_test__name__icontains=search_query)
        )
    if start_date:
        lab_qs = lab_qs.filter(generated_date__date__gte=start_date)
    if end_date:
        lab_qs = lab_qs.filter(generated_date__date__lte=end_date)
    if status_filter and status_filter.lower() != 'all':
        lab_qs = lab_qs.filter(status__iexact=status_filter)

    # 4. Unified Billing Reports Queryset (OPD Bills + IPD Bills + Laboratory Bills)
    opd_bills_qs = OPDVisit.objects.select_related('patient').order_by('-created_at')
    ipd_bills_qs = IPDBill.objects.select_related('patient').order_by('-created_at')
    lab_bills_qs = LaboratoryBill.objects.select_related('patient').order_by('-created_at')

    if start_date:
        opd_bills_qs = opd_bills_qs.filter(visit_date__gte=start_date)
        ipd_bills_qs = ipd_bills_qs.filter(created_at__date__gte=start_date)
        lab_bills_qs = lab_bills_qs.filter(created_at__date__gte=start_date)
    if end_date:
        opd_bills_qs = opd_bills_qs.filter(visit_date__lte=end_date)
        ipd_bills_qs = ipd_bills_qs.filter(created_at__date__lte=end_date)
        lab_bills_qs = lab_bills_qs.filter(created_at__date__lte=end_date)

    hospital_settings = HospitalSettings.objects.first()
    default_consult_fee = hospital_settings.consultation_fee if (hospital_settings and hospital_settings.consultation_fee) else Decimal('200.00')

    billing_list = []

    # OPD Bills
    if billing_type_filter.lower() in ['all', 'opd']:
        for v in opd_bills_qs:
            st = 'Paid' if v.status != 'Cancelled' else 'Cancelled'
            if status_filter and status_filter.lower() != 'all' and status_filter.lower() not in st.lower():
                continue
            fee = default_consult_fee if v.visit_type == OPDVisit.VisitTypeChoices.NEW_VISIT else Decimal('0.00')
            billing_list.append({
                'id': v.id,
                'patient_id': v.patient.id if v.patient else None,
                'bill_number': v.opd_number,
                'patient_name': v.patient.full_name if v.patient else "N/A",
                'patient_uhid': v.patient.uhid if v.patient else "N/A",
                'bill_type': 'OPD Consultation',
                'amount': fee,
                'status': st,
                'date': v.visit_date,
                'created_at': v.created_at,
                'is_opd': True,
                'is_ipd': False,
                'is_lab': False,
            })

    # IPD Bills
    if billing_type_filter.lower() in ['all', 'ipd']:
        for b in ipd_bills_qs:
            st = b.get_payment_status_display() if hasattr(b, 'get_payment_status_display') else str(b.payment_status)
            if status_filter and status_filter.lower() != 'all' and status_filter.lower() not in st.lower():
                continue
            billing_list.append({
                'id': b.id,
                'patient_id': b.patient.id if b.patient else None,
                'bill_number': b.bill_number,
                'patient_name': b.patient.full_name if b.patient else "N/A",
                'patient_uhid': b.patient.uhid if b.patient else "N/A",
                'bill_type': 'IPD Settlement',
                'amount': b.net_amount,
                'status': st,
                'date': b.bill_date,
                'created_at': b.created_at,
                'is_opd': False,
                'is_ipd': True,
                'is_lab': False,
            })

    # Laboratory Bills
    if billing_type_filter.lower() in ['all', 'lab', 'laboratory']:
        for b in lab_bills_qs:
            st = 'Paid'
            if status_filter and status_filter.lower() != 'all' and 'paid' not in status_filter.lower():
                continue
            billing_list.append({
                'id': b.id,
                'patient_id': b.patient.id if b.patient else None,
                'bill_number': b.bill_number,
                'patient_name': b.patient.full_name if b.patient else "N/A",
                'patient_uhid': b.patient.uhid if b.patient else "N/A",
                'bill_type': 'Laboratory Investigation',
                'amount': b.grand_total,
                'status': st,
                'date': b.bill_date,
                'created_at': b.created_at,
                'is_opd': False,
                'is_ipd': False,
                'is_lab': True,
            })

    billing_list.sort(key=lambda x: x['created_at'], reverse=True)

    if search_query:
        sq = search_query.lower()
        billing_list = [
            b for b in billing_list
            if sq in b['bill_number'].lower() or sq in str(b['patient_name']).lower() or sq in str(b['patient_uhid']).lower() or sq in str(b['bill_type']).lower()
        ]

    # Handle Exports if export parameter is supplied
    if export_format in ["excel", "csv"]:
        if active_tab == "opd":
            headers = ["OPD Number", "Patient Name", "UHID", "Doctor", "Visit Date", "Status"]
            data_rows = []
            for v in opd_qs:
                u = v.created_by
                doc_name = f"Dr. {u.first_name} {u.last_name}".strip() if (u and (u.first_name or u.last_name)) else (f"Dr. {u.username}" if u else "N/A")
                v_date = v.visit_date.strftime("%d %b %Y") if v.visit_date else "-"
                data_rows.append([v.opd_number, v.patient.full_name if v.patient else "N/A", v.patient.uhid if v.patient else "N/A", doc_name, v_date, v.get_status_display() or v.status])

        elif active_tab == "ipd":
            headers = ["Admission ID", "Patient Name", "UHID", "Ward Type", "Admission Date", "Discharge Date", "Status"]
            data_rows = []
            for adm in ipd_qs:
                adm_date = adm.admission_date.strftime("%d %b %Y") if adm.admission_date else "-"
                dis_date = adm.discharge_date.strftime("%d %b %Y") if adm.discharge_date else "-"
                data_rows.append([adm.receipt_number or "IPD-ADM", adm.patient.full_name if adm.patient else "N/A", adm.patient.uhid if adm.patient else "N/A", adm.ward_type or "-", adm_date, dis_date, adm.status])

        elif active_tab == "lab":
            headers = ["Report Number", "Patient Name", "UHID", "Test Name", "Report Date", "Status", "Doctor Name"]
            data_rows = []
            for rep in lab_qs:
                rep_date = rep.generated_date.strftime("%d %b %Y") if rep.generated_date else "-"
                u = rep.visit.created_by if rep.visit else None
                doc_name = f"Dr. {u.first_name} {u.last_name}".strip() if (u and (u.first_name or u.last_name)) else (f"Dr. {u.username}" if u else "N/A")
                data_rows.append([rep.report_number, rep.patient.full_name if rep.patient else "N/A", rep.patient.uhid if rep.patient else "N/A", rep.lab_test.name if rep.lab_test else "N/A", rep_date, rep.get_status_display() or rep.status, doc_name])

        elif active_tab == "billing":
            headers = ["Bill Number", "Patient Name", "UHID", "Bill Type", "Amount (INR)", "Payment Status", "Bill Date"]
            data_rows = []
            for b in billing_list:
                b_date = b['date'].strftime("%d %b %Y") if hasattr(b['date'], 'strftime') else str(b['date'])
                data_rows.append([b['bill_number'], b['patient_name'], b['patient_uhid'], b['bill_type'], b['amount'], b['status'], b_date])

        if export_format == "csv":
            import codecs
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = f'attachment; filename="{active_tab}_reports.csv"'
            response.write(codecs.BOM_UTF8)
            writer = csv.writer(response)
            writer.writerow(headers)
            for row in data_rows:
                writer.writerow(row)
            return response

        elif export_format == "excel":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{active_tab.upper()} Reports"

            ws.append(headers)
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in data_rows:
                ws.append(row)
                row_idx = ws.max_row
                for col_num in range(1, len(row) + 1):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.border = thin_border

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{active_tab}_reports.xlsx"'
            wb.save(response)
            return response

    # Paginators
    opd_paginator = Paginator(opd_qs, 10)
    opd_page = opd_paginator.get_page(request.GET.get("page") if active_tab == "opd" else 1)

    ipd_paginator = Paginator(ipd_qs, 10)
    ipd_page = ipd_paginator.get_page(request.GET.get("page") if active_tab == "ipd" else 1)

    lab_paginator = Paginator(lab_qs, 10)
    lab_page = lab_paginator.get_page(request.GET.get("page") if active_tab == "lab" else 1)

    billing_paginator = Paginator(billing_list, 10)
    billing_page = billing_paginator.get_page(request.GET.get("page") if active_tab == "billing" else 1)

    context = {
        'total_opd_reports': f"{opd_reports_count:,}",
        'total_ipd_reports': f"{ipd_reports_count:,}",
        'total_lab_reports': f"{lab_reports_count:,}",
        'total_billing_reports': f"{billing_reports_count:,}",
        'active_tab': active_tab,
        'search_query': search_query,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'status_filter': status_filter,
        'patient_type_filter': patient_type_filter,
        'billing_type_filter': billing_type_filter,
        'opd_page': opd_page,
        'ipd_page': ipd_page,
        'lab_page': lab_page,
        'billing_page': billing_page,
    }
    return render(request, "adminpanel/reports.html", context)


def coming_soon(request, title="Feature Coming Soon"):
    return render(request, "adminpanel/coming_soon.html", {"page_title": title})
